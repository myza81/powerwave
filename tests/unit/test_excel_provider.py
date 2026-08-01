"""Unit tests for app.providers.excel.excel_provider.ExcelProvider."""
from __future__ import annotations

import textwrap
from datetime import datetime
from pathlib import Path

import numpy as np
import openpyxl
import pytest

from app.models import DisturbanceRecord
from app.providers.base.exceptions import ProviderLoadError
from app.providers.excel.excel_provider import (
    ExcelProvider,
    _detect_time_column,
    _estimate_rate,
    _infer_unit,
    _is_digital_column,
    _score_sheet,
    _select_sheet,
)

import pandas as pd


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────


def _make_xlsx(tmp_path: Path, name: str, rows: list[list]) -> Path:
    """Create a single-sheet xlsx workbook from a list-of-rows (first row = headers)."""
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(str(p))
    return p


def _make_multi_sheet_xlsx(
    tmp_path: Path, name: str, sheets: dict[str, list[list]]
) -> Path:
    """Create a multi-sheet xlsx workbook. sheets = {sheet_name: rows}."""
    p = tmp_path / name
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    wb.save(str(p))
    return p


# ─────────────────────────────────────────────────────────────────────────────
# TestCanLoad
# ─────────────────────────────────────────────────────────────────────────────


class TestCanLoad:
    def test_xlsx_returns_true(self) -> None:
        assert ExcelProvider().can_load(Path("data.xlsx")) is True

    def test_xls_returns_true(self) -> None:
        assert ExcelProvider().can_load(Path("data.xls")) is True

    def test_xlsx_uppercase_returns_true(self) -> None:
        assert ExcelProvider().can_load(Path("DATA.XLSX")) is True

    def test_csv_returns_false(self) -> None:
        assert ExcelProvider().can_load(Path("data.csv")) is False

    def test_txt_returns_false(self) -> None:
        assert ExcelProvider().can_load(Path("data.txt")) is False

    def test_no_suffix_returns_false(self) -> None:
        assert ExcelProvider().can_load(Path("data")) is False


# ─────────────────────────────────────────────────────────────────────────────
# TestDetectTimeColumn
# ─────────────────────────────────────────────────────────────────────────────


class TestDetectTimeColumn:
    def test_time_lowercase(self) -> None:
        assert _detect_time_column(["time", "VA", "VB"]) == "time"

    def test_time_uppercase(self) -> None:
        assert _detect_time_column(["Time", "VA"]) == "Time"

    def test_t_single_letter(self) -> None:
        assert _detect_time_column(["t", "VA"]) == "t"

    def test_seconds(self) -> None:
        assert _detect_time_column(["seconds", "IA"]) == "seconds"

    def test_timestamp(self) -> None:
        assert _detect_time_column(["timestamp", "VA"]) == "timestamp"

    def test_datetime(self) -> None:
        assert _detect_time_column(["datetime", "VA"]) == "datetime"

    def test_no_time_column(self) -> None:
        assert _detect_time_column(["VA", "VB", "VC"]) is None

    def test_first_match_wins(self) -> None:
        assert _detect_time_column(["seconds", "time", "VA"]) == "seconds"


# ─────────────────────────────────────────────────────────────────────────────
# TestInferUnit
# ─────────────────────────────────────────────────────────────────────────────


class TestInferUnit:
    def test_voltage_keyword_volt(self) -> None:
        assert _infer_unit("Voltage_A") == "kV"

    def test_voltage_keyword_kv(self) -> None:
        assert _infer_unit("VA_kV") == "kV"

    def test_voltage_short_va(self) -> None:
        assert _infer_unit("va") == "kV"

    def test_current_keyword_curr(self) -> None:
        assert _infer_unit("Current_A") == "A"

    def test_current_short_ia(self) -> None:
        assert _infer_unit("ia") == "A"

    def test_frequency(self) -> None:
        assert _infer_unit("Frequency") == "Hz"

    def test_hz(self) -> None:
        assert _infer_unit("freq_hz") == "Hz"

    def test_mvar(self) -> None:
        assert _infer_unit("Q_MVAr") == "MVar"

    def test_mw(self) -> None:
        assert _infer_unit("Active_MW") == "MW"

    def test_unknown(self) -> None:
        assert _infer_unit("Channel_42") == "unknown"


# ─────────────────────────────────────────────────────────────────────────────
# TestIsDigitalColumn
# ─────────────────────────────────────────────────────────────────────────────


class TestIsDigitalColumn:
    def test_bool_dtype_always_digital(self) -> None:
        s = pd.Series([True, False, True])
        assert _is_digital_column(s, "anything") is True

    def test_binary_with_status_keyword(self) -> None:
        s = pd.Series([0, 1, 0, 1])
        assert _is_digital_column(s, "CB_status") is True

    def test_binary_without_keyword_is_analog(self) -> None:
        s = pd.Series([0, 1, 0, 1])
        assert _is_digital_column(s, "VA") is False

    def test_non_binary_with_keyword_is_analog(self) -> None:
        s = pd.Series([0.0, 1.0, 2.0])
        assert _is_digital_column(s, "trip_voltage") is False

    def test_all_nan_returns_false(self) -> None:
        s = pd.Series([float("nan"), float("nan")])
        assert _is_digital_column(s, "status") is False

    def test_trip_keyword(self) -> None:
        s = pd.Series([0, 0, 1, 0])
        assert _is_digital_column(s, "Trip_signal") is True

    def test_alarm_keyword(self) -> None:
        s = pd.Series([0, 1])
        assert _is_digital_column(s, "alarm") is True


# ─────────────────────────────────────────────────────────────────────────────
# TestEstimateRate
# ─────────────────────────────────────────────────────────────────────────────


class TestEstimateRate:
    def test_uniform_50hz(self) -> None:
        t = np.arange(0, 1.0, 0.02)
        assert _estimate_rate(t) == pytest.approx(50.0, rel=1e-3)

    def test_uniform_1000hz(self) -> None:
        t = np.arange(0, 0.1, 0.001)
        assert _estimate_rate(t) == pytest.approx(1000.0, rel=1e-3)

    def test_single_sample_returns_zero(self) -> None:
        assert _estimate_rate(np.array([0.0])) == 0.0

    def test_empty_returns_zero(self) -> None:
        assert _estimate_rate(np.array([])) == 0.0


# ─────────────────────────────────────────────────────────────────────────────
# TestScoreSheet
# ─────────────────────────────────────────────────────────────────────────────


class TestScoreSheet:
    def test_empty_df_scores_zero(self) -> None:
        assert _score_sheet(pd.DataFrame()) == 0

    def test_numeric_columns_score_positive(self) -> None:
        df = pd.DataFrame({"A": [1.0, 2.0, 3.0], "B": [4.0, 5.0, 6.0]})
        assert _score_sheet(df) > 0

    def test_all_text_columns_score_zero_or_low(self) -> None:
        df = pd.DataFrame({"A": ["x", "y", "z"], "B": ["a", "b", "c"]})
        # text columns that are not numeric-coercible score 0
        assert _score_sheet(df) == 0


# ─────────────────────────────────────────────────────────────────────────────
# TestSelectSheet
# ─────────────────────────────────────────────────────────────────────────────


class TestSelectSheet:
    def test_single_sheet_returned(self, tmp_path: Path) -> None:
        p = _make_xlsx(tmp_path, "s.xlsx", [["VA", "VB"], [1.0, 2.0]])
        assert _select_sheet(p) == "Sheet"

    def test_most_data_rich_sheet_selected(self, tmp_path: Path) -> None:
        p = _make_multi_sheet_xlsx(
            tmp_path,
            "m.xlsx",
            {
                "Small": [["VA"], [1.0]],
                "Data": [
                    ["time", "VA", "VB", "VC", "IA"],
                    [0.0, 1.0, 1.0, 1.0, 0.5],
                    [0.02, 1.1, 1.1, 1.1, 0.6],
                    [0.04, 1.2, 1.2, 1.2, 0.7],
                ],
            },
        )
        assert _select_sheet(p) == "Data"

    def test_missing_file_raises_provider_load_error(self, tmp_path: Path) -> None:
        with pytest.raises(ProviderLoadError, match="Cannot open"):
            _select_sheet(tmp_path / "ghost.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# TestLoadNumericTime
# ─────────────────────────────────────────────────────────────────────────────


class TestLoadNumericTime:
    def test_returns_disturbance_record(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path,
            "n.xlsx",
            [["time", "VA", "VB"], [0.0, 1.0, -1.0], [0.02, 1.1, -1.1]],
        )
        rec = ExcelProvider().load(p)
        assert isinstance(rec, DisturbanceRecord)

    def test_analog_count(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path,
            "n.xlsx",
            [["time", "VA", "VB", "VC"], [0.0, 1.0, 1.0, 1.0], [0.02, 2.0, 2.0, 2.0]],
        )
        rec = ExcelProvider().load(p)
        assert len(rec.analog_channels) == 3

    def test_time_column_excluded_from_channels(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path,
            "n.xlsx",
            [["time", "VA"], [0.0, 1.0], [0.02, 1.1]],
        )
        rec = ExcelProvider().load(p)
        names = [ch.name for ch in rec.analog_channels]
        assert "time" not in names

    def test_sample_rate_inferred(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path,
            "n.xlsx",
            [["time", "VA"]] + [[i * 0.02, float(i)] for i in range(50)],
        )
        rec = ExcelProvider().load(p)
        assert rec.sampling_info.sampling_rates[0] == pytest.approx(50.0, rel=1e-2)

    def test_waveform_data_shape(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path,
            "n.xlsx",
            [["time", "VA", "VB"]] + [[i * 0.001, float(i), float(-i)] for i in range(10)],
        )
        rec = ExcelProvider().load(p)
        assert rec.waveform_data.shape == (10, 3)  # time + VA + VB

    def test_validate_passes(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path,
            "n.xlsx",
            [["time", "VA"]] + [[i * 0.02, float(i)] for i in range(10)],
        )
        rec = ExcelProvider().load(p)
        errors = rec.validate()
        assert errors == []


# ─────────────────────────────────────────────────────────────────────────────
# TestAmbiguousDateOrder — Powerwave policy: day-first default for CSV/Excel
# ─────────────────────────────────────────────────────────────────────────────


class TestAmbiguousDateOrder:
    def test_ambiguous_date_defaults_day_first(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path,
            "ambiguous.xlsx",
            [["time", "VA"], ["3/6/2026 17:25", 1.0], ["3/6/2026 17:26", 1.1]],
        )
        rec = ExcelProvider().load(p)
        assert rec.timing_info.start_time == datetime(2026, 6, 3, 17, 25, 0)

    def test_ambiguous_date_emits_warning(self, tmp_path: Path) -> None:
        import warnings

        p = _make_xlsx(
            tmp_path,
            "ambiguous.xlsx",
            [["time", "VA"], ["3/6/2026 17:25", 1.0], ["3/6/2026 17:26", 1.1]],
        )
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            ExcelProvider().load(p)
        text = "\n".join(str(w.message) for w in caught)
        assert "ambiguous date" in text.lower()
        assert "DD/MM/YYYY" in text

    def test_unambiguous_day_over_twelve_no_warning(self, tmp_path: Path) -> None:
        import warnings

        p = _make_xlsx(
            tmp_path,
            "unambiguous.xlsx",
            [["time", "VA"], ["13/6/2026 17:25", 1.0], ["13/6/2026 17:26", 1.1]],
        )
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            rec = ExcelProvider().load(p)
        text = "\n".join(str(w.message) for w in caught)
        assert "ambiguous date" not in text.lower()
        assert rec.timing_info.start_time == datetime(2026, 6, 13, 17, 25, 0)


# ─────────────────────────────────────────────────────────────────────────────
# TestLoadNoTimeColumn
# ─────────────────────────────────────────────────────────────────────────────


class TestLoadNoTimeColumn:
    def test_integer_index_used(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path,
            "no_time.xlsx",
            [["VA", "VB"], [1.0, -1.0], [2.0, -2.0], [3.0, -3.0]],
        )
        rec = ExcelProvider().load(p)
        assert rec.waveform_data["time"].tolist() == [0.0, 1.0, 2.0]

    def test_sample_rate_zero(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path,
            "no_time.xlsx",
            [["VA", "VB"], [1.0, 2.0], [3.0, 4.0]],
        )
        rec = ExcelProvider().load(p)
        assert rec.sampling_info.sampling_rates[0] == 0.0

    def test_epoch_fallback_start_time(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path,
            "no_time.xlsx",
            [["VA"], [1.0], [2.0]],
        )
        rec = ExcelProvider().load(p)
        assert rec.timing_info.start_time == datetime(2000, 1, 1)


# ─────────────────────────────────────────────────────────────────────────────
# TestUnitInference
# ─────────────────────────────────────────────────────────────────────────────


class TestUnitInference:
    def test_voltage_column_unit(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path, "u.xlsx",
            [["time", "VA_kV"], [0.0, 275.0], [0.02, 276.0]],
        )
        rec = ExcelProvider().load(p)
        assert rec.analog_channels[0].unit == "kV"

    def test_current_column_unit(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path, "u.xlsx",
            [["time", "ia"], [0.0, 1.0], [0.02, 1.1]],
        )
        rec = ExcelProvider().load(p)
        assert rec.analog_channels[0].unit == "A"

    def test_frequency_column_unit(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path, "u.xlsx",
            [["time", "frequency"], [0.0, 50.0], [0.02, 50.0]],
        )
        rec = ExcelProvider().load(p)
        assert rec.analog_channels[0].unit == "Hz"

    def test_unknown_column_unit(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path, "u.xlsx",
            [["time", "channel_x"], [0.0, 1.0], [0.02, 2.0]],
        )
        rec = ExcelProvider().load(p)
        assert rec.analog_channels[0].unit == "unknown"


# ─────────────────────────────────────────────────────────────────────────────
# TestParameterTypeInference — shared semantic classifier integration
# ─────────────────────────────────────────────────────────────────────────────


class TestParameterTypeInference:
    def test_confident_demand_column_populates_active_power(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path, "d.xlsx",
            [["time", "System Demand"]] + [[i * 0.02, 18700.0 + i] for i in range(5)],
        )
        rec = ExcelProvider().load(p)
        ch = next(c for c in rec.analog_channels if c.name == "System Demand")
        assert ch.name == "System Demand"
        assert ch.unit == "MW"
        assert ch.parameter_type == "active_power"

    def test_low_confidence_tie_line_leaves_unit_and_parameter_type_unset(self, tmp_path: Path) -> None:
        # Same confidence discipline now applies to unit as already applied
        # to parameter_type -- "Tie-Line" is name-derived but below the
        # confirmation threshold, so neither field is populated.
        p = _make_xlsx(
            tmp_path, "d.xlsx",
            [["time", "Tie-Line"]] + [[i * 0.02, 80.0 + i] for i in range(5)],
        )
        rec = ExcelProvider().load(p)
        ch = next(c for c in rec.analog_channels if c.name == "Tie-Line")
        assert ch.unit == "unknown"
        assert ch.parameter_type is None

    def test_unrecognised_column_leaves_parameter_type_none(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path, "u.xlsx",
            [["time", "channel_x"], [0.0, 1.0], [0.02, 2.0]],
        )
        rec = ExcelProvider().load(p)
        assert rec.analog_channels[0].parameter_type is None

    # -- Value-only magnitude fallback: unit and parameter_type both gated ---

    def test_neutral_header_near_1pu_does_not_populate_unit_or_type(self, tmp_path: Path) -> None:
        import numpy as np

        rng = np.random.default_rng(42)
        vals = (1.0 + rng.normal(0, 0.02, 30)).tolist()
        p = _make_xlsx(
            tmp_path, "n.xlsx",
            [["time", "Column 1"]] + [[i * 0.02, v] for i, v in enumerate(vals)],
        )
        rec = ExcelProvider().load(p)
        ch = next(c for c in rec.analog_channels if c.name == "Column 1")
        assert ch.unit == "unknown"
        assert ch.parameter_type is None

    def test_neutral_header_noisy_large_magnitude_does_not_populate_unit_or_type(self, tmp_path: Path) -> None:
        import numpy as np

        rng = np.random.default_rng(7)
        vals = (18700.0 + rng.normal(0, 15.0, 30)).tolist()
        p = _make_xlsx(
            tmp_path, "n.xlsx",
            [["time", "Column 1"]] + [[i * 0.02, v] for i, v in enumerate(vals)],
        )
        rec = ExcelProvider().load(p)
        ch = next(c for c in rec.analog_channels if c.name == "Column 1")
        assert ch.unit == "unknown"
        assert ch.parameter_type is None

    def test_neutral_header_negative_power_like_does_not_populate_unit_or_type(self, tmp_path: Path) -> None:
        import numpy as np

        rng = np.random.default_rng(7)
        vals = (-120.0 + rng.normal(0, 8.0, 30)).tolist()
        p = _make_xlsx(
            tmp_path, "n.xlsx",
            [["time", "Column 1"]] + [[i * 0.02, v] for i, v in enumerate(vals)],
        )
        rec = ExcelProvider().load(p)
        ch = next(c for c in rec.analog_channels if c.name == "Column 1")
        assert ch.unit == "unknown"
        assert ch.parameter_type is None


# ─────────────────────────────────────────────────────────────────────────────
# TestDigitalInference
# ─────────────────────────────────────────────────────────────────────────────


class TestDigitalInference:
    def test_binary_status_column_is_digital(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path, "d.xlsx",
            [["time", "VA", "CB_status"], [0.0, 1.0, 0], [0.02, 1.1, 1]],
        )
        rec = ExcelProvider().load(p)
        assert len(rec.digital_channels) == 1
        assert rec.digital_channels[0].name == "CB_status"

    def test_binary_without_keyword_is_analog(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path, "d.xlsx",
            [["time", "VA", "VB_binary"], [0.0, 1.0, 0], [0.02, 1.1, 1]],
        )
        rec = ExcelProvider().load(p)
        assert len(rec.digital_channels) == 0
        assert len(rec.analog_channels) == 2

    def test_digital_channel_index_sequential(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path, "d.xlsx",
            [
                ["time", "trip_A", "trip_B"],
                [0.0, 0, 0],
                [0.02, 1, 0],
            ],
        )
        rec = ExcelProvider().load(p)
        assert len(rec.digital_channels) == 2
        assert rec.digital_channels[0].index == 0
        assert rec.digital_channels[1].index == 1


# ─────────────────────────────────────────────────────────────────────────────
# TestMultiSheetSelection
# ─────────────────────────────────────────────────────────────────────────────


class TestMultiSheetSelection:
    def test_loads_most_data_rich_sheet(self, tmp_path: Path) -> None:
        p = _make_multi_sheet_xlsx(
            tmp_path,
            "multi.xlsx",
            {
                "Config": [["Key", "Value"], ["rate", "50"]],
                "Waveform": [
                    ["time", "VA", "VB", "VC"],
                    [0.0, 275.0, -137.0, -138.0],
                    [0.02, 274.0, -137.5, -136.5],
                    [0.04, 273.0, -137.0, -136.0],
                ],
            },
        )
        rec = ExcelProvider().load(p)
        assert len(rec.analog_channels) == 3  # VA, VB, VC from Waveform sheet

    def test_single_sheet_workbook_loads_correctly(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path, "single.xlsx",
            [["time", "VA"], [0.0, 1.0], [0.02, 1.1]],
        )
        rec = ExcelProvider().load(p)
        assert isinstance(rec, DisturbanceRecord)


# ─────────────────────────────────────────────────────────────────────────────
# TestErrorHandling
# ─────────────────────────────────────────────────────────────────────────────


class TestErrorHandling:
    def test_missing_file_raises(self, tmp_path: Path) -> None:
        with pytest.raises(ProviderLoadError, match="not found"):
            ExcelProvider().load(tmp_path / "ghost.xlsx")

    def test_xls_raises_clear_message(self, tmp_path: Path) -> None:
        fake_xls = tmp_path / "data.xls"
        fake_xls.write_bytes(b"fake")
        with pytest.raises(ProviderLoadError, match="xlrd"):
            ExcelProvider().load(fake_xls)

    def test_empty_sheet_raises(self, tmp_path: Path) -> None:
        p = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.save(str(p))
        with pytest.raises(ProviderLoadError):
            ExcelProvider().load(p)

    def test_no_usable_columns_raises(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path, "bad.xlsx",
            [["time", "label"], [0.0, "abc"], [0.02, "def"]],
        )
        with pytest.raises(ProviderLoadError, match="no usable"):
            ExcelProvider().load(p)


# ─────────────────────────────────────────────────────────────────────────────
# TestChannelIndexing
# ─────────────────────────────────────────────────────────────────────────────


class TestChannelIndexing:
    def test_analog_indices_sequential(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path, "idx.xlsx",
            [["time", "VA", "VB", "VC"], [0.0, 1.0, 1.0, 1.0], [0.02, 2.0, 2.0, 2.0]],
        )
        rec = ExcelProvider().load(p)
        assert [ch.index for ch in rec.analog_channels] == [0, 1, 2]

    def test_mixed_channels_indexed_independently(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path, "mixed.xlsx",
            [
                ["time", "VA", "trip_A", "VB"],
                [0.0, 1.0, 0, -1.0],
                [0.02, 1.1, 1, -1.1],
            ],
        )
        rec = ExcelProvider().load(p)
        assert len(rec.analog_channels) == 2
        assert len(rec.digital_channels) == 1
        assert rec.analog_channels[0].index == 0
        assert rec.analog_channels[1].index == 1
        assert rec.digital_channels[0].index == 0


# ─────────────────────────────────────────────────────────────────────────────
# TestMetadata
# ─────────────────────────────────────────────────────────────────────────────


class TestMetadata:
    def test_provider_type_is_excel(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path, "m.xlsx",
            [["time", "VA"], [0.0, 1.0], [0.02, 1.1]],
        )
        rec = ExcelProvider().load(p)
        assert rec.metadata.provider_type == "excel"

    def test_source_file_recorded(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path, "m.xlsx",
            [["time", "VA"], [0.0, 1.0], [0.02, 1.1]],
        )
        rec = ExcelProvider().load(p)
        assert str(p) in rec.metadata.source_file

    def test_recorder_name_is_excel(self, tmp_path: Path) -> None:
        p = _make_xlsx(
            tmp_path, "m.xlsx",
            [["time", "VA"], [0.0, 1.0], [0.02, 1.1]],
        )
        rec = ExcelProvider().load(p)
        assert rec.metadata.recorder_name == "Excel"

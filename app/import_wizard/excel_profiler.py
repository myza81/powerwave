"""Excel file profiling using openpyxl (read-only, no GUI/pandas dependency).

Functions
---------
get_sheet_names(path)           List sheet names without loading cell data.
profile_excel(path, ...)        Build RawPreviewModel from one sheet.
"""
from __future__ import annotations

from typing import Any

from app.import_wizard.contracts import ValidationMessage, ValidationSeverity
from app.import_wizard.models import RawPreviewModel


def get_sheet_names(path: str) -> list[str]:
    """Return the sheet names from an Excel workbook.

    Returns an empty list on any error (missing file, not Excel, etc.).
    """
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception:
        return []


def _cell_value(cell: Any) -> Any:
    """Extract a clean Python value from an openpyxl cell."""
    val = cell.value
    if val is None:
        return ""
    if isinstance(val, str):
        return val.strip()
    return val


def _is_mostly_numeric_values(cells: list[Any], threshold: float = 0.5) -> bool:
    non_empty = [c for c in cells if c != "" and c is not None]
    if not non_empty:
        return False
    numeric = 0
    for cell in non_empty:
        if isinstance(cell, (int, float)):
            numeric += 1
            continue
        try:
            float(str(cell).strip())
            numeric += 1
        except ValueError:
            continue
    return (numeric / len(non_empty)) >= threshold


def _alpha_score_values(cells: list[Any]) -> float:
    non_empty = [c for c in cells if c != "" and c is not None]
    if not non_empty:
        return 0.0
    alpha = sum(
        1
        for c in non_empty
        if isinstance(c, str) and sum(ch.isalpha() for ch in c) / max(len(c), 1) >= 0.5
    )
    return alpha / len(non_empty)


def _looks_like_time_header(value: Any) -> bool:
    text = str(value).strip().lower()
    return text in {
        "time",
        "t",
        "sec",
        "secs",
        "seconds",
        "timestamp",
        "datetime",
        "elapsed",
        "duration",
    }


def _find_header_row_index_excel(rows: list[list[Any]]) -> int:
    """Lookahead header detection for Excel rows (values, not strings)."""
    for i, row in enumerate(rows):
        non_empty = [c for c in row if c != "" and c is not None]
        if len(non_empty) < 2:
            continue
        alpha_score = _alpha_score_values(row)
        has_time_header = any(_looks_like_time_header(c) for c in non_empty)
        if alpha_score < 0.40 and not has_time_header:
            continue
        for j in range(i + 1, len(rows)):
            next_row = rows[j]
            if not any(c != "" and c is not None for c in next_row):
                continue
            if _is_mostly_numeric_values(next_row):
                return i
            break

    for i, row in enumerate(rows):
        if any(c != "" and c is not None for c in row):
            return i
    return 0


def profile_excel(
    path: str,
    *,
    sheet_name: str | None = None,
    max_preview_rows: int = 50,
    max_scan_rows: int = 200,
) -> tuple[RawPreviewModel, list[ValidationMessage]]:
    """Build a RawPreviewModel for one sheet of an Excel workbook.

    Parameters
    ----------
    path:               Absolute path to the .xlsx / .xls file.
    sheet_name:         Sheet to profile; defaults to the first sheet.
    max_preview_rows:   Maximum data rows in the preview.
    max_scan_rows:      Maximum rows to read when scanning for the header.

    Returns
    -------
    (model, warnings)
    """
    warnings: list[ValidationMessage] = []

    try:
        import openpyxl
    except ImportError:
        return (
            RawPreviewModel(column_names=[], preview_rows=[]),
            [
                ValidationMessage(
                    severity=ValidationSeverity.ERROR,
                    code="EXCEL_OPENPYXL_MISSING",
                    message="openpyxl is not installed; Excel files cannot be profiled.",
                )
            ],
        )

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return (
            RawPreviewModel(column_names=[], preview_rows=[]),
            [
                ValidationMessage(
                    severity=ValidationSeverity.ERROR,
                    code="EXCEL_OPEN_FAILED",
                    message=f"Could not open Excel file: {exc}",
                )
            ],
        )

    available_sheets = list(wb.sheetnames)
    if not available_sheets:
        wb.close()
        return (
            RawPreviewModel(column_names=[], preview_rows=[]),
            [
                ValidationMessage(
                    severity=ValidationSeverity.ERROR,
                    code="EXCEL_NO_SHEETS",
                    message="Workbook contains no sheets.",
                )
            ],
        )

    active_sheet = sheet_name or available_sheets[0]
    if active_sheet not in available_sheets:
        warnings.append(
            ValidationMessage(
                severity=ValidationSeverity.WARNING,
                code="EXCEL_SHEET_NOT_FOUND",
                message=f"Sheet '{active_sheet}' not found; using '{available_sheets[0]}'.",
            )
        )
        active_sheet = available_sheets[0]

    ws = wb[active_sheet]

    # Read up to max_scan_rows rows
    scan_rows: list[list[Any]] = []
    for row in ws.iter_rows(max_row=max_scan_rows):
        scan_rows.append([_cell_value(cell) for cell in row])

    wb.close()

    if not scan_rows:
        return (
            RawPreviewModel(column_names=[], preview_rows=[], sheet_name=active_sheet),
            [
                ValidationMessage(
                    severity=ValidationSeverity.ERROR,
                    code="EXCEL_EMPTY_SHEET",
                    message=f"Sheet '{active_sheet}' is empty.",
                )
            ],
        )

    header_idx = _find_header_row_index_excel(scan_rows)
    skipped_count = header_idx

    if header_idx >= len(scan_rows):
        return (
            RawPreviewModel(column_names=[], preview_rows=[], sheet_name=active_sheet),
            [
                ValidationMessage(
                    severity=ValidationSeverity.ERROR,
                    code="EXCEL_NO_HEADER",
                    message="Could not locate a header row.",
                )
            ],
        )

    header_row = scan_rows[header_idx]
    column_names_raw = [str(c) if c != "" and c is not None else "" for c in header_row]

    # Deduplicate blank/duplicate names
    seen: dict[str, int] = {}
    column_names: list[str] = []
    for name in column_names_raw:
        key = name.strip() or "Column"
        if key in seen:
            seen[key] += 1
            column_names.append(f"{key}_{seen[key]}")
        else:
            seen[key] = 0
            column_names.append(key.strip() if key != "Column" else key)

    expected_cols = len(column_names)
    data_rows: list[list[Any]] = []
    for row in scan_rows[header_idx + 1 : header_idx + 1 + max_preview_rows]:
        if not any(c != "" and c is not None for c in row):
            continue
        row_vals = list(row)
        if len(row_vals) < expected_cols:
            row_vals += [""] * (expected_cols - len(row_vals))
        elif len(row_vals) > expected_cols:
            row_vals = row_vals[:expected_cols]
        data_rows.append(row_vals)

    while column_names and column_names[-1] == "Column" and all(
        (len(row) <= len(column_names) - 1 or row[-1] in ("", None))
        for row in data_rows
    ):
        column_names.pop()
        column_names_raw.pop()
        for row in data_rows:
            if row:
                row.pop()

    # Row count estimate: use ws.max_row minus header rows (may be None for read-only)
    try:
        wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws2 = wb2[active_sheet]
        max_row = ws2.max_row or 0
        wb2.close()
        row_count_estimate = max(0, max_row - header_idx - 1)
    except Exception:
        row_count_estimate = len(data_rows)

    return (
        RawPreviewModel(
            column_names=column_names,
            preview_rows=data_rows,
            header_row_index=header_idx,
            skipped_row_count=skipped_count,
            row_count_estimate=row_count_estimate,
            sheet_name=active_sheet,
            parse_warnings=[w.message for w in warnings],
        ),
        warnings,
    )
